import sys
from googlesearch import search
import os
import time
import requests
import re
import tkinter as tk
from tkinter import ttk, simpledialog, Menu
from tkinter import *
from html.parser import HTMLParser
from bs4 import BeautifulSoup
import ctypes
import getpass
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import string

def printtext():
    global e
    global string
    string = e.get()
    text.insert(INSERT, string)

URL_DATA = []

#setting up user input box
root = tk.Tk()
root.withdraw()
root.title('Name')
text = Text(root)
e = Entry(root)
e.pack()
e.focus_set()
b = Button(root, text='okay', command=printtext)
text.pack()
b.pack(side='bottom')

#search prompt (ex. academicaffairs.kennesaw.edu)
SRC_INP = simpledialog.askstring(title="Prompt", text="\nThank you for using Sam's Grocery Thingy!\n", prompt="Would you like to search items individually or from your Alexa list?")

#search logic
x = 0
while x == 0:
    if SRC_INP == None:
        print("No valid entry was detected, please try again.")
        SRC_INP
    elif SRC_INP == '' or SRC_INP.find('.') == -1:
        sys.exit("This is not a valid URL. Please restart the program and try something else.")
    else:
        print("Searching stores for your items...")

#appending URLs to system
for url in search(('site:'+SRC_INP), tld='com', lang='en'):
    URL_DATA.append(url)

if ('urllib.error.HTTPError: HTTP Error 429: Too Many Requests'):
    time.sleep(2)


print("URLs indexed to be scraped:",URL_DATA,'\n','\n','\n')

time.sleep(2000)

print("Scraping indexed URLs for metadata...",'\n')

#printing data
for url in URL_DATA:
    if url == None:
        continue    
    else:
        if url.find('.pdf') == -1:
            response = requests.get(url)
            page = response.content
            soup = BeautifulSoup(page, 'lxml')
            print(url)
            URL_DATA.append(soup.find('div', { 'class' : 'content' }).find('p', recursive=False))
            print("Entry complete",'\n')
        else:
            pass

print("Search complete!")

#creating Excel workbook
wb = Workbook()

dest_filename = 'URL_list.xlsx'

ws1 = wb.active
ws1.title = "Site Data"

wb.save(filename='URL_list.xlsx')

#saving to Excel
for col in range(1,2):
    ws1.append(URL_DATA)

root = tk.Tk()
root.withdraw()
root.title('Name')

XLSX_INP = simpledialog.askstring(title="Excel Workbook Name", prompt="What would you like to name the Excel sheet for this data?")
XLSX_INP

while XLSX_INP == None:
    print("Error in saving workbook, please retry.")
    XLSX_INP

else:
    print("Saving to Excel...")
    wb.save(filename = XLSX_INP+'.xlsx')
    sys.exit("Thank you for flying Air Sam!")
